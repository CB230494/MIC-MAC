import io
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(page_title="MIC MAC desde plantilla Excel", layout="wide")

AZUL = "#1F5F8B"
VERDE = "#1B9E77"
NARANJA = "#E67E22"
ROJO = "#CC3D3D"
GRIS = "#6B7280"
TEXTO = "#124559"

plt.rcParams.update({
    "figure.dpi": 180,
    "savefig.dpi": 180,
    "axes.titlesize": 16,
    "axes.labelsize": 12,
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
    "axes.grid": True,
    "grid.alpha": 0.25,
})


# -----------------------------------------------------------------------------
# Lectura de plantilla
# -----------------------------------------------------------------------------
def limpiar_texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def deduplicar_lista(items: List[str]) -> List[str]:
    salida = []
    vistos = set()
    for item in items:
        limpio = limpiar_texto(item)
        if not limpio:
            continue
        clave = limpio.casefold()
        if clave not in vistos:
            vistos.add(clave)
            salida.append(limpio)
    return salida


def cargar_diccionario_descriptores(ws) -> pd.DataFrame:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_corto = limpiar_texto(row[0])
        descriptor = limpiar_texto(row[1])
        categoria = limpiar_texto(row[2])
        descripcion = limpiar_texto(row[3])
        if nombre_corto:
            rows.append({
                "codigo": nombre_corto,
                "descriptor": descriptor or nombre_corto,
                "categoria": categoria,
                "descripcion": descripcion,
            })
    df = pd.DataFrame(rows).drop_duplicates(subset=["codigo"], keep="first")
    return df


def extraer_instituciones(ws_matriz) -> Tuple[pd.DataFrame, pd.DataFrame]:
    participantes = []
    for row in range(9, 26):
        participante = limpiar_texto(ws_matriz.cell(row, 1).value)
        institucion = limpiar_texto(ws_matriz.cell(row, 5).value)
        puesto = limpiar_texto(ws_matriz.cell(row, 9).value)
        if participante or institucion or puesto:
            participantes.append({
                "participante": participante,
                "institucion": institucion,
                "puesto": puesto,
            })

    df_participantes = pd.DataFrame(participantes)
    instituciones = deduplicar_lista(df_participantes.get("institucion", pd.Series(dtype=str)).tolist())
    df_instituciones = pd.DataFrame({"institucion": instituciones})
    return df_participantes, df_instituciones


def detectar_bloque_matriz(ws_matriz) -> Tuple[int, int, int]:
    fila_encabezados = None
    col_inicio = None
    col_fin = None

    for r in range(1, ws_matriz.max_row + 1):
        valores = [limpiar_texto(ws_matriz.cell(r, c).value) for c in range(1, ws_matriz.max_column + 1)]
        no_vacios = [i + 1 for i, v in enumerate(valores) if v]
        if len(no_vacios) >= 5:
            primer_valor = valores[no_vacios[0] - 1]
            if primer_valor == "":
                fila_encabezados = r
                col_inicio = no_vacios[0]
                col_fin = no_vacios[-1]
                break

    if fila_encabezados is None:
        raise ValueError("No se pudo detectar la fila de encabezados de la matriz MIC MAC.")

    return fila_encabezados, col_inicio, col_fin


def extraer_matriz(ws_matriz) -> Tuple[pd.DataFrame, List[str]]:
    fila_encabezados, col_inicio, col_fin = detectar_bloque_matriz(ws_matriz)
    codigos_columnas = [
        limpiar_texto(ws_matriz.cell(fila_encabezados, c).value)
        for c in range(col_inicio, col_fin + 1)
    ]
    codigos_columnas = [c for c in codigos_columnas if c]

    fila_inicio_datos = fila_encabezados + 1
    matriz = []
    codigos_filas = []

    for r in range(fila_inicio_datos, ws_matriz.max_row + 1):
        codigo_fila = limpiar_texto(ws_matriz.cell(r, col_inicio - 1).value)
        if not codigo_fila:
            if codigos_filas:
                break
            continue

        valores = [ws_matriz.cell(r, c).value for c in range(col_inicio, col_fin + 1)]
        valores_limpios = pd.to_numeric(pd.Series(valores), errors="coerce").fillna(0).astype(int).tolist()
        matriz.append(valores_limpios[:len(codigos_columnas)])
        codigos_filas.append(codigo_fila)

    df = pd.DataFrame(matriz, index=codigos_filas, columns=codigos_columnas)
    df = df.apply(pd.to_numeric, errors="coerce").fillna(0).astype(int).clip(lower=0, upper=3)

    comun = [c for c in df.columns if c in df.index]
    df = df.loc[comun, comun].copy()

    for i in range(len(df)):
        df.iat[i, i] = 0

    return df, comun


def extraer_frecuencias(ws_descriptores) -> pd.DataFrame:
    encabezados = [limpiar_texto(c.value).upper() for c in ws_descriptores[1]]
    idx_codigo = 0 if encabezados else None
    idx_freq = None

    posibles = {
        "FRECUENCIA", "FRECUENCIAS", "FREQ", "TOTAL", "CANTIDAD", "CONTEO", "%", "PORCENTAJE"
    }

    for i, nombre in enumerate(encabezados):
        if nombre in posibles:
            idx_freq = i
            break

    if idx_codigo is None or idx_freq is None:
        return pd.DataFrame(columns=["codigo", "frecuencia"])

    rows = []
    for row in ws_descriptores.iter_rows(min_row=2, values_only=True):
        codigo = limpiar_texto(row[idx_codigo])
        frecuencia = pd.to_numeric(row[idx_freq], errors="coerce")
        if codigo and pd.notna(frecuencia):
            rows.append({"codigo": codigo, "frecuencia": float(frecuencia)})

    return pd.DataFrame(rows)


def leer_plantilla_excel(file) -> Dict[str, pd.DataFrame]:
    wb = openpyxl.load_workbook(file, data_only=True)

    hoja_descriptores = None
    hoja_matriz = None
    for nombre in wb.sheetnames:
        nombre_norm = limpiar_texto(nombre).upper()
        if "DESCRIPTOR" in nombre_norm:
            hoja_descriptores = wb[nombre]
        if "MATRIZ" in nombre_norm:
            hoja_matriz = wb[nombre]

    if hoja_descriptores is None or hoja_matriz is None:
        raise ValueError("El archivo debe incluir una hoja de descriptores y una hoja de matriz.")

    df_dicc = cargar_diccionario_descriptores(hoja_descriptores)
    df_participantes, df_instituciones = extraer_instituciones(hoja_matriz)
    df_matriz, codigos = extraer_matriz(hoja_matriz)
    df_freq = extraer_frecuencias(hoja_descriptores)

    return {
        "diccionario": df_dicc,
        "participantes": df_participantes,
        "instituciones": df_instituciones,
        "matriz": df_matriz,
        "frecuencias": df_freq,
        "codigos": pd.DataFrame({"codigo": codigos}),
    }


# -----------------------------------------------------------------------------
# Cálculo MIC MAC
# -----------------------------------------------------------------------------
def clasificar_zona(influencia: float, dependencia: float, media_inf: float, media_dep: float) -> str:
    if influencia == 0 and dependencia == 0:
        return "Sin relación"
    if influencia >= media_inf and dependencia >= media_dep:
        return "Conflicto"
    if influencia >= media_inf and dependencia < media_dep:
        return "Poder"
    if influencia < media_inf and dependencia >= media_dep:
        return "Resultado"
    return "Autónoma"



def analizar_micmac(df_matriz: pd.DataFrame, df_diccionario: pd.DataFrame, df_freq: pd.DataFrame) -> pd.DataFrame:
    influencia = df_matriz.sum(axis=1)
    dependencia = df_matriz.sum(axis=0)

    df = pd.DataFrame({
        "codigo": df_matriz.index,
        "influencia": influencia.values,
        "dependencia": dependencia.values,
    })

    media_inf = df["influencia"].mean() if not df.empty else 0
    media_dep = df["dependencia"].mean() if not df.empty else 0
    df["zona"] = df.apply(
        lambda r: clasificar_zona(r["influencia"], r["dependencia"], media_inf, media_dep),
        axis=1,
    )

    df = df.merge(df_diccionario, on="codigo", how="left")
    if not df_freq.empty:
        df = df.merge(df_freq, on="codigo", how="left")
    else:
        df["frecuencia"] = np.nan

    orden_zona = {
        "Poder": 1,
        "Conflicto": 2,
        "Resultado": 3,
        "Autónoma": 4,
        "Sin relación": 5,
    }
    df["orden_zona"] = df["zona"].map(orden_zona).fillna(99)
    df = df.sort_values(["orden_zona", "influencia", "dependencia", "descriptor"], ascending=[True, False, False, True]).reset_index(drop=True)
    df["frecuencia"] = pd.to_numeric(df["frecuencia"], errors="coerce")
    return df


# -----------------------------------------------------------------------------
# Visualizaciones
# -----------------------------------------------------------------------------
def mapa_micmac_png(df: pd.DataFrame, titulo: str) -> bytes:
    fig, ax = plt.subplots(figsize=(12, 8))

    colores = {
        "Conflicto": ROJO,
        "Poder": VERDE,
        "Resultado": AZUL,
        "Autónoma": NARANJA,
        "Sin relación": GRIS,
    }

    if df.empty:
        ax.axis("off")
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center", fontsize=16)
    else:
        x = df["dependencia"].to_numpy(dtype=float)
        y = df["influencia"].to_numpy(dtype=float)
        mean_x = x.mean() if len(x) else 0
        mean_y = y.mean() if len(y) else 0

        ax.axvline(mean_x, linestyle="--", color="#666666", linewidth=1)
        ax.axhline(mean_y, linestyle="--", color="#666666", linewidth=1)

        offsets = np.linspace(-0.15, 0.15, len(df)) if len(df) > 1 else np.array([0.0])

        for i, (_, row) in enumerate(df.iterrows()):
            xx = row["dependencia"] + offsets[i]
            yy = row["influencia"] + offsets[::-1][i]
            ax.scatter(
                xx,
                yy,
                s=150,
                color=colores.get(row["zona"], AZUL),
                edgecolors="black",
                linewidths=0.6,
                alpha=0.9,
            )
            etiqueta = row["descriptor"] if pd.notna(row["descriptor"]) and row["descriptor"] else row["codigo"]
            ax.text(xx + 0.18, yy + 0.18, etiqueta, fontsize=8, ha="left", va="bottom")

        ax.set_xlim(left=min(x.min() - 1, 0), right=x.max() + 2)
        ax.set_ylim(bottom=min(y.min() - 1, 0), top=y.max() + 2)
        ax.set_xlabel("Dependencia")
        ax.set_ylabel("Influencia")
        ax.set_title(titulo, color=TEXTO)

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.08)
    plt.close(fig)
    return buf.getvalue()


# -----------------------------------------------------------------------------
# Excel de salida
# -----------------------------------------------------------------------------
def exportar_excel_resultados(nombre_fuente: str, data: Dict[str, pd.DataFrame], df_resultados: pd.DataFrame) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        wb = writer.book

        fmt_header = wb.add_format({
            "bold": True,
            "bg_color": "#0B3954",
            "font_color": "white",
            "border": 0,
            "text_wrap": True,
            "align": "center",
            "valign": "vcenter",
        })
        fmt_sub = wb.add_format({
            "bold": True,
            "bg_color": "#DCEAF6",
        })
        fmt_num = wb.add_format({"align": "center"})

        resumen = pd.DataFrame({
            "campo": [
                "Archivo fuente",
                "Cantidad de problemáticas",
                "Cantidad de instituciones únicas",
                "Participantes registrados",
                "Promedio influencia",
                "Promedio dependencia",
            ],
            "valor": [
                nombre_fuente,
                len(df_resultados),
                len(data["instituciones"]),
                len(data["participantes"]),
                round(df_resultados["influencia"].mean(), 2) if not df_resultados.empty else 0,
                round(df_resultados["dependencia"].mean(), 2) if not df_resultados.empty else 0,
            ]
        })
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        ws = writer.sheets["Resumen"]
        ws.set_column("A:A", 30)
        ws.set_column("B:B", 30)
        for c, col in enumerate(resumen.columns):
            ws.write(0, c, col, fmt_header)

        data["instituciones"].to_excel(writer, sheet_name="Instituciones", index=False)
        ws = writer.sheets["Instituciones"]
        ws.set_column("A:A", 55)
        ws.write(0, 0, "institucion", fmt_header)

        data["participantes"].to_excel(writer, sheet_name="Participantes", index=False)
        ws = writer.sheets["Participantes"]
        ws.set_column("A:A", 32)
        ws.set_column("B:B", 45)
        ws.set_column("C:C", 30)
        for c, col in enumerate(data["participantes"].columns):
            ws.write(0, c, col, fmt_header)

        df_matriz_export = data["matriz"].copy()
        df_matriz_export.insert(0, "codigo", df_matriz_export.index)
        df_matriz_export.to_excel(writer, sheet_name="Matriz_Limpia", index=False)
        ws = writer.sheets["Matriz_Limpia"]
        ws.set_column(0, 0, 16)
        ws.set_column(1, len(df_matriz_export.columns) - 1, 10, fmt_num)
        for c, col in enumerate(df_matriz_export.columns):
            ws.write(0, c, col, fmt_header)

        cols_res = ["codigo", "descriptor", "categoria", "frecuencia", "influencia", "dependencia", "zona"]
        df_resultados[cols_res].to_excel(writer, sheet_name="Analisis_MICMAC", index=False)
        ws = writer.sheets["Analisis_MICMAC"]
        ws.set_column("A:A", 14)
        ws.set_column("B:B", 55)
        ws.set_column("C:C", 22)
        ws.set_column("D:F", 14, fmt_num)
        ws.set_column("G:G", 16)
        for c, col in enumerate(cols_res):
            ws.write(0, c, col, fmt_header)

        zonas_orden = ["Poder", "Conflicto", "Resultado", "Autónoma", "Sin relación"]
        filas = []
        for zona in zonas_orden:
            grupo = df_resultados[df_resultados["zona"] == zona].copy()
            if grupo.empty:
                continue
            for _, row in grupo.iterrows():
                filas.append({
                    "zona": zona,
                    "codigo": row["codigo"],
                    "descriptor": row["descriptor"],
                    "categoria": row["categoria"],
                    "frecuencia": row["frecuencia"],
                    "influencia": row["influencia"],
                    "dependencia": row["dependencia"],
                })
        pd.DataFrame(filas).to_excel(writer, sheet_name="Problematicas_por_zona", index=False)
        ws = writer.sheets["Problematicas_por_zona"]
        ws.set_column("A:A", 16)
        ws.set_column("B:B", 14)
        ws.set_column("C:C", 55)
        ws.set_column("D:D", 22)
        ws.set_column("E:G", 14, fmt_num)
        for c, col in enumerate(["zona", "codigo", "descriptor", "categoria", "frecuencia", "influencia", "dependencia"]):
            ws.write(0, c, col, fmt_header)

    return output.getvalue()


# -----------------------------------------------------------------------------
# Interfaz
# -----------------------------------------------------------------------------
st.title("MIC MAC independiente desde plantilla Excel")
st.caption("Sube la plantilla MIC MAC. La app leerá cruces, depurará instituciones duplicadas y generará el análisis por zona.")

archivo = st.file_uploader("Carga la plantilla MIC MAC en Excel", type=["xlsx"])

if archivo is not None:
    try:
        data = leer_plantilla_excel(archivo)
        df_resultados = analizar_micmac(data["matriz"], data["diccionario"], data["frecuencias"])

        st.success("Plantilla procesada correctamente.")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Problemáticas", len(df_resultados))
        with col2:
            st.metric("Instituciones únicas", len(data["instituciones"]))
        with col3:
            st.metric("Participantes", len(data["participantes"]))
        with col4:
            st.metric("Cruces totales", int(data["matriz"].to_numpy().sum()))

        st.divider()
        st.subheader("Instituciones depuradas")
        st.dataframe(data["instituciones"], use_container_width=True, hide_index=True)

        st.subheader("Participantes registrados")
        st.dataframe(data["participantes"], use_container_width=True, hide_index=True)

        st.subheader("Resultados MIC MAC")
        mostrar = df_resultados[["codigo", "descriptor", "categoria", "frecuencia", "influencia", "dependencia", "zona"]].copy()
        st.dataframe(mostrar, use_container_width=True, hide_index=True)

        st.subheader("Ubicación de las problemáticas por zona")
        zonas = ["Poder", "Conflicto", "Resultado", "Autónoma", "Sin relación"]
        tabs = st.tabs(zonas)
        for tab, zona in zip(tabs, zonas):
            with tab:
                grupo = mostrar[mostrar["zona"] == zona].copy()
                if grupo.empty:
                    st.info(f"No hay problemáticas ubicadas en {zona}.")
                else:
                    st.dataframe(grupo, use_container_width=True, hide_index=True)

        st.subheader("Mapa MIC MAC")
        png = mapa_micmac_png(df_resultados, "Mapa MIC MAC")
        st.image(png, use_container_width=True)

        excel_salida = exportar_excel_resultados(Path(archivo.name).name, data, df_resultados)
        st.download_button(
            "Descargar resultados en Excel",
            data=excel_salida,
            file_name=f"Resultados_MICMAC_{Path(archivo.name).stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.download_button(
            "Descargar mapa MIC MAC (PNG)",
            data=png,
            file_name=f"Mapa_MICMAC_{Path(archivo.name).stem}.png",
            mime="image/png",
            use_container_width=True,
        )

    except Exception as e:
        st.error(f"No se pudo procesar la plantilla: {e}")
else:
    st.info("Sube una plantilla para comenzar.")






